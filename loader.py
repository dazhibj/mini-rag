from pathlib import Path
from typing import Callable, Generator

import openpyxl


def load_text_file(path: Path) -> str | None:
    try:
        return path.read_text(encoding="utf-8")
    except Exception as e:
        print(f"Error reading {path}: {e}")
        return None


def load_excel_file(path: Path) -> str | None:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"Error reading {path}: {e}")
        return None
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            rows.append(" | ".join(cells))
        parts.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts)


SUPPORTED_EXTS: dict[str, Callable[[Path], str | None]] = {
    ".txt": load_text_file,
    ".md": load_text_file,
    ".xlsx": load_excel_file,
    ".xls": load_excel_file,
}


def load_documents(directory: Path) -> Generator[tuple[str, str, str], None, None]:
    for path in directory.rglob("*"):
        if not path.is_file():
            continue
        loader = SUPPORTED_EXTS.get(path.suffix.lower())
        if loader is None:
            continue
        content = loader(path)
        if content:
            rel = path.relative_to(directory)
            yield str(rel), str(path.suffix), content
